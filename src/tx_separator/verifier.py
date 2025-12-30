"""Verification logic to ensure output matches input."""

import csv
from pathlib import Path
from typing import List, Tuple

import xlrd
from openpyxl import load_workbook


def detect_file_type(path: Path) -> str:
    """Detect file type based on extension."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".xls":
        return "xls"
    return "csv"


def get_sorted_output_files(output_dir: Path, output_format: str) -> List[Path]:
    """Get all transaction output files sorted by name."""
    ext = "xlsx" if output_format == "xlsx" else "csv"
    return sorted(output_dir.glob(f"transactions_*.{ext}"))


def read_csv_rows(file_path: Path) -> List[List[str]]:
    """Read all rows from a tab-delimited CSV file."""
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter="\t")
        return list(reader)


def read_xlsx_rows(file_path: Path) -> List[List[str]]:
    """Read all rows from an Excel .xlsx file."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return rows


def read_xls_rows(file_path: Path) -> List[List[str]]:
    """Read all rows from an Excel .xls file (legacy format)."""
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    rows = []
    for row_idx in range(ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell_value(row_idx, col_idx)
            # Convert all values to strings
            if isinstance(cell, float) and cell == int(cell):
                row.append(str(int(cell)))
            else:
                row.append(str(cell) if cell != "" else "")
        rows.append(row)
    return rows


def read_file_rows(file_path: Path) -> List[List[str]]:
    """Read all rows from a file (CSV or Excel)."""
    file_type = detect_file_type(file_path)
    if file_type == "xlsx":
        return read_xlsx_rows(file_path)
    if file_type == "xls":
        return read_xls_rows(file_path)
    return read_csv_rows(file_path)


def verify_output(
    input_file: Path,
    output_dir: Path,
    has_header: bool = False,
    output_format: str = "auto",
) -> Tuple[bool, str]:
    """
    Verify that concatenated output files match the original input.

    When has_header=True, accounts for the fact that the header appears once
    in the input but is repeated in each output file.

    Args:
        input_file: Path to the original input file
        output_dir: Directory containing output files
        has_header: Whether files have a header row
        output_format: Output format used ('csv', 'xlsx', or 'auto')

    Returns:
        Tuple of (success: bool, message: str)
    """
    # Detect formats (xls input defaults to xlsx output)
    input_type = detect_file_type(input_file)
    if output_format == "auto":
        output_format = "xlsx" if input_type in ("xlsx", "xls") else "csv"

    # Read original input
    original_rows = read_file_rows(input_file)

    # Extract header from original if present
    original_header = None
    if has_header and original_rows:
        original_header = original_rows[0]
        original_rows = original_rows[1:]

    # Read and concatenate all output files (sorted by filename)
    output_files = get_sorted_output_files(output_dir, output_format)

    if not output_files:
        return False, "No output files found"

    concatenated_rows = []
    for output_file in output_files:
        file_rows = read_file_rows(output_file)

        if has_header and file_rows:
            # Verify header matches in each output file
            output_header = file_rows[0]
            if original_header and output_header != original_header:
                return False, (
                    f"Header mismatch in {output_file.name}: "
                    f"expected {original_header}, got {output_header}"
                )
            # Skip header when concatenating
            file_rows = file_rows[1:]

        concatenated_rows.extend(file_rows)

    # Compare row counts
    if len(original_rows) != len(concatenated_rows):
        return False, (
            f"Row count mismatch: original has {len(original_rows)} data rows, "
            f"output has {len(concatenated_rows)} data rows"
        )

    # Check each row
    for i, (orig, out) in enumerate(zip(original_rows, concatenated_rows)):
        if orig != out:
            return False, f"Row {i + 1} differs between original and output"

    return True, "Verification successful: output matches original"
