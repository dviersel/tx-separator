"""Core transaction separation logic."""

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import xlrd
from openpyxl import Workbook, load_workbook


def parse_date(date_str: str) -> datetime:
    """Parse date string in YYYYMMDD format."""
    return datetime.strptime(str(date_str).split(".")[0], "%Y%m%d")


def get_output_filename(date: datetime, output_format: str) -> str:
    """Generate output filename for a given date."""
    ext = "xlsx" if output_format == "xlsx" else "csv"
    return f"transactions_{date.strftime('%Y_%m')}.{ext}"


def detect_file_type(path: Path) -> str:
    """Detect file type based on extension."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".xls":
        return "xls"
    return "csv"


def read_csv_rows(input_file: Path) -> List[List[str]]:
    """Read all rows from a tab-delimited CSV file."""
    with open(input_file, "r", encoding="utf-8") as file:
        reader = csv.reader(file, delimiter="\t")
        return [row for row in reader]


def read_xlsx_rows(input_file: Path) -> List[List[str]]:
    """Read all rows from an Excel .xlsx file."""
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Convert all values to strings, handling None
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return rows


def read_xls_rows(input_file: Path) -> List[List[str]]:
    """Read all rows from an Excel .xls file (legacy format)."""
    wb = xlrd.open_workbook(input_file)
    ws = wb.sheet_by_index(0)
    rows = []
    for row_idx in range(ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell_value(row_idx, col_idx)
            # Convert all values to strings
            if isinstance(cell, float) and cell == int(cell):
                row.append(str(int(cell)))
            else:
                row.append(str(cell) if cell != "" else "")
        rows.append(row)
    return rows


def write_csv_file(
    output_path: Path, rows: List[List[str]], header: List[str] | None = None
) -> None:
    """Write rows to a tab-delimited CSV file."""
    with open(output_path, "w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file, delimiter="\t")
        if header:
            writer.writerow(header)
        writer.writerows(rows)


def write_xlsx_file(
    output_path: Path, rows: List[List[str]], header: List[str] | None = None
) -> None:
    """Write rows to an Excel file."""
    wb = Workbook()
    ws = wb.active
    if header:
        ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(output_path)


def process_transactions(
    input_file: Path,
    output_dir: Path,
    has_header: bool = False,
    output_format: str = "auto",
) -> Tuple[Dict[str, int], List[str] | None]:
    """
    Process transactions from input file and split by month.

    Args:
        input_file: Path to the input file (CSV or Excel)
        output_dir: Directory to write output files
        has_header: Whether the input file has a header row
        output_format: Output format - 'csv', 'xlsx', or 'auto' (match input)

    Returns:
        Tuple of:
        - Dictionary mapping output filenames to transaction counts
        - Header row (if has_header=True) or None
    """
    # Detect input file type
    input_type = detect_file_type(input_file)

    # Determine output format (xls input defaults to xlsx output)
    if output_format == "auto":
        output_format = "xlsx" if input_type in ("xlsx", "xls") else "csv"

    # Read input file
    if input_type == "xlsx":
        all_rows = read_xlsx_rows(input_file)
    elif input_type == "xls":
        all_rows = read_xls_rows(input_file)
    else:
        all_rows = read_csv_rows(input_file)

    # Extract header if present
    header = None
    if has_header and all_rows:
        header = all_rows[0]
        all_rows = all_rows[1:]

    # Group transactions by month
    monthly_transactions: Dict[str, List[List[str]]] = {}

    for row in all_rows:
        # Date is in the 3rd column (index 2)
        transaction_date = parse_date(row[2])
        output_file = get_output_filename(transaction_date, output_format)

        if output_file not in monthly_transactions:
            monthly_transactions[output_file] = []

        monthly_transactions[output_file].append(row)

    # Write transactions to separate files
    file_counts = {}
    for output_file, transactions in monthly_transactions.items():
        full_output_path = output_dir / output_file

        if output_format == "xlsx":
            write_xlsx_file(full_output_path, transactions, header)
        else:
            write_csv_file(full_output_path, transactions, header)

        file_counts[output_file] = len(transactions)

    return file_counts, header
